from openpyxl import load_workbook

wb = load_workbook("Libro1.xlsx")

ws = wb["Hoja1"]

cells = ws["C4:F10"]

mat = []

for r in cells:
    row = []
    for cell in r:
        x = cell.value
        row.append(x)
    mat.append(row)

# Una vez recuperada la matriz "mat"
# vamos a transformarla en una lista
# de diccionarios

keys = mat[0]

personas = []

for i in range(1, len(mat)):
    persona = {}

    for j in range(0, len(keys)):
        k = keys[j]

        persona[k] = mat[i][j]

    personas.append(persona)

print personas