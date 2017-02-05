from openpyxl import load_workbook

# 1. Cargar el libro de excel (workbook)
wb = load_workbook("Libro1.xlsx")

# 2. Recuperar del libro la hoja (worksheet)
#ws = wb["Hoja 1"]
ws = wb.active

# 3. Obtenemos de la hoja el rango de celdas
cells = ws["C4:F10"]

# 4. Recorrer las filas del rango de celdas
for row in cells:
    for cell in row:
        print cell.value

    print "-" * 30

