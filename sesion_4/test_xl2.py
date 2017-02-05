from openpyxl import load_workbook

wb = load_workbook("Libro1.xlsx")

ws = wb["Hoja1"]

cells = ws["C4:F10"]

mat = []

for r in cells:
    row = []
    for cell in r:
        x = cell.value
        row.append(x)
    mat.append(row)

print mat

x = ws["C4"].value

x = mat[0][0]