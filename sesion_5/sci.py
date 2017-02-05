from openpyxl import load_workbook

def range_xl(filename, sheet_name, cell_range):
    wb = load_workbook(filename)
    ws = wb[sheet_name]
    cells = ws[cell_range]
    return cells

def matrix_xl(filename, sheet_name, cell_range):
    cells = range_xl(filename, sheet_name, cell_range)
    mat = []
    for row_xl in cells:
        row = []
        for cell in row_xl:
            row.append(cell.value)
        mat.append(row)
    return mat

def dic_xl(filename, sheet_name, cell_range):
    mat = matrix_xl(filename, sheet_name, cell_range)
    keys = mat[0]
    coll = []
    for i in range(1, len(mat)):
        dic = {}
        for j in range(0, len(keys)):
            k = keys[j]
            dic[k] = mat[i][j]
        coll.append(dic)
    return coll


