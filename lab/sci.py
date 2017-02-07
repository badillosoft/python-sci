#
# Alan Badillo Salas <badillo.soft@hotmail.com>
# https://github.com/badillosoft
#
# Python's Scientist Module
#

from openpyxl import load_workbook

def clean_init_spaces(s):
	aux = ""
	i = 0
	while i < len(s) and s[i] == " ":
		i += 1
	while i < len(s):
		aux += s[i]
		i += 1
	return aux

def clean_end_spaces(s):
	aux = ""
	i = 0
	while i < len(s) and s[i] != " " and s[i] != "\n":
		aux += s[i]
		i += 1
	return aux

def range_xl(filename, sheet_name, cell_range):
    wb = load_workbook(filename)
    ws = wb[sheet_name]
    cells = ws[cell_range]
    return cells


def matrix_xl(cells):
    mat = []
    for row_xl in cells:
        row = []
        for cell in row_xl:
            row.append(cell.value)
        mat.append(row)
    return mat


def matrix_csv(filename):
	f = open(filename, "r")
	mat = []
	for line in f:
		row_csv = line.split(",")
		row = []
		for cell in row_csv:
			try:
				row.append(float(cell))
			except:
				s = clean_init_spaces(cell)
				s = clean_end_spaces(s)
				row.append(s)
		mat.append(row)
	f.close()
	return mat


def data_build(mat):
    keys = mat[0]
    coll = []
    for i in range(1, len(mat)):
        dic = {}
        for j in range(0, len(keys)):
			k = keys[j]
			if k != None and k != "":
				dic[k] = mat[i][j]
        coll.append(dic)
    return coll

def load_xl(filename, sheet_name, cell_range):
	cells = range_xl(filename, sheet_name, cell_range)
	mat = matrix_xl(cells)
	return data_build(mat)

def load_csv(filename):
	mat = matrix_csv(filename)
	return data_build(mat)

def data_extract(data, key):
	aux = []
	for dic in data:
		aux.append(dic[key])
	return aux

def data_filter(data, fn_filter):
	aux = []
	for dic in data:
		if fn_filter(dic):
			aux.append(dic)
	return aux

def data_checker(data, fn_checker):
	aux = []
	for dic in data:
		aux.append(fn_checker(dic))
	return aux

def data_reduce(data, reduce):
	aux = []
	for i in range(min(len(data), len(reduce))):
		if (reduce[i]):
			aux.append(data[i])
	return aux

def data_map(data, fn_map):
	aux = []
	for dic in data:
		x = fn_map(dic)
		if x != None:
			aux.append(x)
	return aux
